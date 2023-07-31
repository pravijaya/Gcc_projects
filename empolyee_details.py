import openpyxl

def add_employee_data():
    employees = []

    while True:
        print("Enter employee details:")
        name = input("Name: ")
        employee_id = input("employee_id: ")
        salary = input("salary:")
        phone_no = input("Phone No: ")

        employee = {
            "Name": name,
            "employee_id": employee_id,
            "salary": salary,
            "Phone No": phone_no,
        }
        employees.append(employee)

        more_data = input("Do you want to enter data for another employee? (yes/no): ")
        if more_data.lower() != "yes":
            break

    return employees

def create_excel(employees, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(employees[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Write student data
    for row_idx, employee in enumerate(employees, start=2):
        for col_idx, value in enumerate(employee.values(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the Excel file
    workbook.save(output_file)

if __name__ == "__main__":
    print("Enter student data. Type 'done' when you finish.")

    employee_data = add_employee_data()

    if employee_data:
        print("\nemployee data collected successfully.")
        output_file_name = input("Enter the output Excel file name: ")

        create_excel(employee_data, output_file_name)
        print(f"Data successfully converted and saved to {output_file_name}.")
    else:
        print("No employee data to convert.")